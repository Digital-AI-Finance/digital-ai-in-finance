import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

# Define the project tasks with all details
tasks = [
    # PHASE 1: PRE-WORKSHOP
    {"ID": "1", "Task": "PHASE 1: PRE-WORKSHOP", "Start": "2024-12-01", "End": "2026-03-31", "Type": "Phase", "Resources": ""},

    # Sub-Phase 1.1: Initial Setup
    {"ID": "1.1", "Task": "Initial Setup", "Start": "2024-12-01", "End": "2025-02-28", "Type": "SubPhase", "Resources": ""},
    {"ID": "1.1.1", "Task": "CCG Application Submission", "Start": "2024-12-01", "End": "2024-12-31", "Type": "Task", "Resources": "FHGR Lead"},
    {"ID": "1.1.2", "Task": "Committee Formation", "Start": "2024-12-01", "End": "2025-02-28", "Type": "Task", "Resources": "FHGR/AUS"},
    {"ID": "1.1.3", "Task": "Website Development", "Start": "2024-12-01", "End": "2025-02-28", "Type": "Task", "Resources": "Technical Team"},
    {"ID": "M1", "Task": "MILESTONE: CCG Approval", "Start": "2025-02-28", "End": "2025-02-28", "Type": "Milestone", "Resources": ""},

    # Sub-Phase 1.2: Call for Papers & Marketing
    {"ID": "1.2", "Task": "Call for Papers & Marketing", "Start": "2025-02-01", "End": "2025-08-31", "Type": "SubPhase", "Resources": ""},
    {"ID": "1.2.1", "Task": "Call for Papers Issuance", "Start": "2025-02-01", "End": "2025-03-31", "Type": "Task", "Resources": "Committee"},
    {"ID": "1.2.2", "Task": "Speaker Invitations", "Start": "2025-02-01", "End": "2025-08-31", "Type": "Task", "Resources": "FHGR/AUS"},
    {"ID": "1.2.3", "Task": "Marketing Campaigns", "Start": "2025-03-01", "End": "2025-08-31", "Type": "Task", "Resources": "Marketing Team"},
    {"ID": "1.2.4", "Task": "Paper Submissions", "Start": "2025-03-01", "End": "2025-08-31", "Type": "Task", "Resources": "Researchers"},
    {"ID": "1.2.5", "Task": "Registration Opening", "Start": "2025-03-01", "End": "2025-08-31", "Type": "Task", "Resources": "AUS Team"},

    # Sub-Phase 1.3: Review & Program Development
    {"ID": "1.3", "Task": "Review & Program Development", "Start": "2025-09-01", "End": "2025-10-31", "Type": "SubPhase", "Resources": ""},
    {"ID": "1.3.1", "Task": "Paper Review Process", "Start": "2025-09-01", "End": "2025-10-31", "Type": "Task", "Resources": "Review Committee"},
    {"ID": "1.3.2", "Task": "Acceptance Notifications", "Start": "2025-09-01", "End": "2025-10-31", "Type": "Task", "Resources": "Committee"},
    {"ID": "1.3.3", "Task": "Program Finalization", "Start": "2025-09-01", "End": "2025-10-31", "Type": "Task", "Resources": "FHGR/AUS"},

    # Sub-Phase 1.4: Speaker & Venue Confirmations
    {"ID": "1.4", "Task": "Speaker & Venue Confirmations", "Start": "2025-11-01", "End": "2026-01-31", "Type": "SubPhase", "Resources": ""},
    {"ID": "1.4.1", "Task": "Speaker Confirmations", "Start": "2025-11-01", "End": "2026-01-31", "Type": "Task", "Resources": "FHGR/AUS"},
    {"ID": "1.4.2", "Task": "Venue Planning", "Start": "2025-11-01", "End": "2026-01-31", "Type": "Task", "Resources": "AUS Team"},
    {"ID": "1.4.3", "Task": "Technical Setup Planning", "Start": "2025-11-01", "End": "2026-01-31", "Type": "Task", "Resources": "Technical Team"},

    # Sub-Phase 1.5: Final Preparations
    {"ID": "1.5", "Task": "Final Preparations", "Start": "2026-02-01", "End": "2026-03-31", "Type": "SubPhase", "Resources": ""},
    {"ID": "1.5.1", "Task": "Materials Preparation", "Start": "2026-02-01", "End": "2026-03-31", "Type": "Task", "Resources": "Committee"},
    {"ID": "1.5.2", "Task": "Bilateral Agreement Drafting", "Start": "2026-02-01", "End": "2026-03-31", "Type": "Task", "Resources": "FHGR/AUS Legal"},
    {"ID": "1.5.3", "Task": "Final Logistics", "Start": "2026-02-01", "End": "2026-03-31", "Type": "Task", "Resources": "AUS Team"},
    {"ID": "M2", "Task": "MILESTONE: All Speakers Confirmed", "Start": "2026-03-31", "End": "2026-03-31", "Type": "Milestone", "Resources": ""},

    # PHASE 2: WORKSHOP
    {"ID": "2", "Task": "PHASE 2: WORKSHOP", "Start": "2026-04-01", "End": "2026-04-23", "Type": "Phase", "Resources": ""},

    # Sub-Phase 2.1: Final Preparations
    {"ID": "2.1", "Task": "Final Workshop Preparations", "Start": "2026-04-01", "End": "2026-04-14", "Type": "SubPhase", "Resources": ""},
    {"ID": "2.1.1", "Task": "Final Preparations", "Start": "2026-04-01", "End": "2026-04-14", "Type": "Task", "Resources": "All Teams"},
    {"ID": "2.1.2", "Task": "Technical Testing", "Start": "2026-04-01", "End": "2026-04-14", "Type": "Task", "Resources": "Technical Team"},
    {"ID": "2.1.3", "Task": "Registration Push", "Start": "2026-04-01", "End": "2026-04-14", "Type": "Task", "Resources": "Marketing Team"},
    {"ID": "M3", "Task": "MILESTONE: 80+ Participants Registered", "Start": "2026-04-15", "End": "2026-04-15", "Type": "Milestone", "Resources": ""},

    # Sub-Phase 2.2: Workshop Execution
    {"ID": "2.2", "Task": "Workshop Execution", "Start": "2026-04-21", "End": "2026-04-23", "Type": "SubPhase", "Resources": ""},
    {"ID": "2.2.1", "Task": "Day 1: Research Frontiers", "Start": "2026-04-21", "End": "2026-04-21", "Type": "Task", "Resources": "All Participants"},
    {"ID": "2.2.2", "Task": "Day 2: Industry Applications", "Start": "2026-04-22", "End": "2026-04-22", "Type": "Task", "Resources": "All Participants"},
    {"ID": "2.2.3", "Task": "Day 3: Training & Working Groups", "Start": "2026-04-23", "End": "2026-04-23", "Type": "Task", "Resources": "All Participants"},
    {"ID": "M4", "Task": "MILESTONE: Workshop Delivered", "Start": "2026-04-23", "End": "2026-04-23", "Type": "Milestone", "Resources": ""},
    {"ID": "M5", "Task": "MILESTONE: 15-20 Papers Presented", "Start": "2026-04-23", "End": "2026-04-23", "Type": "Milestone", "Resources": ""},
    {"ID": "2.2.4", "Task": "AUS-FHGR Agreement Signing", "Start": "2026-04-23", "End": "2026-04-23", "Type": "Task", "Resources": "Leadership"},
    {"ID": "M6", "Task": "MILESTONE: Bilateral Agreement Signed", "Start": "2026-04-23", "End": "2026-04-23", "Type": "Milestone", "Resources": ""},

    # PHASE 3: POST-WORKSHOP
    {"ID": "3", "Task": "PHASE 3: POST-WORKSHOP", "Start": "2026-05-01", "End": "2026-07-31", "Type": "Phase", "Resources": ""},

    # Sub-Phase 3.1: Immediate Follow-up
    {"ID": "3.1", "Task": "Immediate Follow-up", "Start": "2026-05-01", "End": "2026-05-31", "Type": "SubPhase", "Resources": ""},
    {"ID": "3.1.1", "Task": "Feedback Collection", "Start": "2026-05-01", "End": "2026-05-31", "Type": "Task", "Resources": "Committee"},
    {"ID": "3.1.2", "Task": "Impact Assessment", "Start": "2026-05-01", "End": "2026-05-31", "Type": "Task", "Resources": "FHGR/AUS"},
    {"ID": "3.1.3", "Task": "Proceedings Compilation", "Start": "2026-05-01", "End": "2026-05-31", "Type": "Task", "Resources": "Editorial Team"},

    # Sub-Phase 3.2: Reporting & Publication
    {"ID": "3.2", "Task": "Reporting & Publication", "Start": "2026-06-01", "End": "2026-06-30", "Type": "SubPhase", "Resources": ""},
    {"ID": "3.2.1", "Task": "CCG Reporting", "Start": "2026-06-01", "End": "2026-06-30", "Type": "Task", "Resources": "FHGR Lead"},
    {"ID": "3.2.2", "Task": "Proceedings Publication", "Start": "2026-06-01", "End": "2026-06-30", "Type": "Task", "Resources": "Editorial Team"},
    {"ID": "M7", "Task": "MILESTONE: Proceedings Published", "Start": "2026-06-30", "End": "2026-06-30", "Type": "Milestone", "Resources": ""},
    {"ID": "3.2.3", "Task": "Media Documentation", "Start": "2026-06-01", "End": "2026-06-30", "Type": "Task", "Resources": "Marketing Team"},

    # Sub-Phase 3.3: Sustainability Initiatives
    {"ID": "3.3", "Task": "Sustainability Initiatives", "Start": "2026-07-01", "End": "2026-07-31", "Type": "SubPhase", "Resources": ""},
    {"ID": "3.3.1", "Task": "Collaboration Platform Launch", "Start": "2026-07-01", "End": "2026-07-31", "Type": "Task", "Resources": "Technical Team"},
    {"ID": "3.3.2", "Task": "Working Group Activation", "Start": "2026-07-01", "End": "2026-07-31", "Type": "Task", "Resources": "Working Groups"},
    {"ID": "3.3.3", "Task": "Proposal Initiation", "Start": "2026-07-01", "End": "2026-07-31", "Type": "Task", "Resources": "FHGR/AUS"},
    {"ID": "M8", "Task": "MILESTONE: Report to CCG Submitted", "Start": "2026-07-31", "End": "2026-07-31", "Type": "Milestone", "Resources": ""}
]

def create_gantt_chart():
    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    phase_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subphase_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    task_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    milestone_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set up headers
    headers = ["ID", "Task Name", "Start Date", "End Date", "Duration", "Resources"]

    # Add month columns from Dec 2024 to Jul 2026
    months = []
    current_date = datetime(2024, 12, 1)
    end_date = datetime(2026, 8, 1)

    while current_date < end_date:
        months.append(current_date.strftime("%b %Y"))
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)

    all_headers = headers + months

    # Write headers
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write tasks
    row = 2
    for task in tasks:
        # Basic task information
        ws.cell(row=row, column=1, value=task["ID"]).border = border

        # Task name with indentation based on level
        task_cell = ws.cell(row=row, column=2, value=task["Task"])
        task_cell.border = border

        # Apply indentation and formatting based on type
        if task["Type"] == "Phase":
            task_cell.font = Font(bold=True, size=11)
            ws.row_dimensions[row].height = 20
        elif task["Type"] == "SubPhase":
            task_cell.value = "  " + task["Task"]
            task_cell.font = Font(bold=True, size=10)
        elif task["Type"] == "Milestone":
            task_cell.value = "    " + task["Task"]
            task_cell.font = Font(bold=True, italic=True, color="FF0000")
        else:
            task_cell.value = "    " + task["Task"]

        # Dates and duration
        start_date = datetime.strptime(task["Start"], "%Y-%m-%d")
        end_date = datetime.strptime(task["End"], "%Y-%m-%d")
        duration = (end_date - start_date).days + 1

        ws.cell(row=row, column=3, value=start_date.strftime("%d/%m/%Y")).border = border
        ws.cell(row=row, column=4, value=end_date.strftime("%d/%m/%Y")).border = border
        ws.cell(row=row, column=5, value=duration).border = border
        ws.cell(row=row, column=6, value=task["Resources"]).border = border

        # Create Gantt bars
        for col, month_str in enumerate(months, len(headers) + 1):
            month_date = datetime.strptime(month_str, "%b %Y")
            month_end = month_date.replace(day=28) if month_date.month == 2 else \
                       month_date.replace(day=30) if month_date.month in [4, 6, 9, 11] else \
                       month_date.replace(day=31)

            cell = ws.cell(row=row, column=col)
            cell.border = border

            # Check if task spans this month
            if start_date <= month_end and end_date >= month_date:
                if task["Type"] == "Phase":
                    cell.fill = phase_fill
                elif task["Type"] == "SubPhase":
                    cell.fill = subphase_fill
                elif task["Type"] == "Task":
                    cell.fill = task_fill
                elif task["Type"] == "Milestone":
                    cell.fill = milestone_fill
                    cell.value = "◆"
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20

    for col in range(len(headers) + 1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Add Resource Summary Sheet
    ws2 = wb.create_sheet("Resource Summary")

    # Resource headers
    resource_headers = ["Resource", "Total Hours", "Contribution (CHF)"]
    for col, header in enumerate(resource_headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Resource data
    resources = [
        ["FHGR Lead", "32 hours", "CHF 2,500 (in-kind)"],
        ["AUS Lead", "50 hours", "CHF 10,500 (in-kind)"],
        ["AUS Cash", "-", "CHF 600"],
        ["CCG Funding", "-", "CHF 5,000"],
        ["", "", ""],
        ["Total Budget", "82 hours", "CHF 18,000"],
        ["Co-funding", "", "CHF 13,000 (72%)"]
    ]

    for i, resource in enumerate(resources, 2):
        for j, value in enumerate(resource, 1):
            cell = ws2.cell(row=i, column=j)
            cell.value = value
            cell.border = border
            if i == 7:  # Total row
                cell.font = Font(bold=True)
            if i == 8:  # Co-funding row
                cell.font = Font(bold=True, italic=True)

    # Adjust column widths
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 25

    # Add Milestones Summary Sheet
    ws3 = wb.create_sheet("Milestones")

    # Milestone headers
    milestone_headers = ["Milestone", "Date", "Description"]
    for col, header in enumerate(milestone_headers, 1):
        cell = ws3.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Milestone data
    milestones = [
        ["CCG Approval", "February 2025", "Funding secured to proceed"],
        ["All Speakers Confirmed", "March 2026", "Speaker lineup finalized"],
        ["80+ Participants Registered", "April 15, 2026", "Registration target achieved"],
        ["Workshop Delivered", "April 21-23, 2026", "3-day workshop executed"],
        ["15-20 Papers Presented", "April 21-23, 2026", "Academic content delivered"],
        ["Bilateral Agreement Signed", "April 23, 2026", "AUS-FHGR collaboration formalized"],
        ["Proceedings Published", "June 2026", "Workshop outputs disseminated"],
        ["Report to CCG Submitted", "July 2026", "Final reporting completed"]
    ]

    for i, milestone in enumerate(milestones, 2):
        for j, value in enumerate(milestone, 1):
            cell = ws3.cell(row=i, column=j)
            cell.value = value
            cell.border = border
            cell.font = Font(bold=(j==1))

    # Adjust column widths
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 40

    # Save the file
    filename = "20241031_CCG_Workshop_Gantt_Chart.xlsx"
    filepath = os.path.join(r"D:\Joerg\Research\slides\2025_LeadingHouse_Travel", filename)
    wb.save(filepath)
    print(f"Gantt chart created successfully: {filepath}")
    return filepath

if __name__ == "__main__":
    excel_file = create_gantt_chart()