import os
import win32com.client as win32
from pathlib import Path

def excel_to_pdf(excel_path, pdf_path):
    """
    Convert Excel file to PDF using Microsoft Excel via COM
    """
    # Get absolute paths
    excel_path = os.path.abspath(excel_path)
    pdf_path = os.path.abspath(pdf_path)

    # Check if Excel file exists
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        return False

    try:
        # Initialize Excel application
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False  # Don't show Excel window
        excel.DisplayAlerts = False  # Don't show alerts

        # Open the workbook
        workbook = excel.Workbooks.Open(excel_path)

        # Configure page setup for better PDF output
        for sheet in workbook.Worksheets:
            # Set to landscape orientation
            sheet.PageSetup.Orientation = 2  # xlLandscape

            # Fit to page
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = False

            # Set margins (in points)
            sheet.PageSetup.LeftMargin = 36  # 0.5 inch
            sheet.PageSetup.RightMargin = 36
            sheet.PageSetup.TopMargin = 36
            sheet.PageSetup.BottomMargin = 36

            # Center on page
            sheet.PageSetup.CenterHorizontally = True

            # Add header with sheet name
            sheet.PageSetup.CenterHeader = f"&B{sheet.Name}&B"

            # Add footer with page number
            sheet.PageSetup.CenterFooter = "Page &P of &N"

        # Export as PDF (0 = xlTypePDF)
        workbook.ExportAsFixedFormat(0, pdf_path)

        # Close workbook and quit Excel
        workbook.Close(SaveChanges=False)
        excel.Quit()

        print(f"Successfully converted to PDF: {pdf_path}")
        return True

    except Exception as e:
        print(f"Error during conversion: {str(e)}")
        try:
            excel.Quit()
        except:
            pass
        return False

# Alternative method using openpyxl and reportlab if COM doesn't work
def excel_to_pdf_alternative(excel_path, pdf_path):
    """
    Alternative method to create PDF representation of Excel data
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from openpyxl import load_workbook

    # Load the Excel file
    wb = load_workbook(excel_path)

    # Create PDF
    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=20
    )

    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Add sheet title
        elements.append(Paragraph(sheet_name, title_style))
        elements.append(Spacer(1, 0.2*inch))

        # Get data from sheet
        data = []
        max_row = min(ws.max_row, 100)  # Limit rows for PDF
        max_col = min(ws.max_column, 15)  # Limit columns for PDF

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                # Truncate long strings
                if isinstance(value, str) and len(value) > 50:
                    value = value[:47] + "..."
                row_data.append(str(value))
            data.append(row_data)

        if data:
            # Create table
            table = Table(data)

            # Style the table
            table_style = TableStyle([
                # Header row
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),

                # Data rows
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),

                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),

                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')])
            ])

            table.setStyle(table_style)
            elements.append(table)

        # Add page break between sheets
        if sheet_name != wb.sheetnames[-1]:
            elements.append(PageBreak())

    # Build PDF
    doc.build(elements)
    print(f"PDF created using alternative method: {pdf_path}")
    return True

if __name__ == "__main__":
    excel_file = r"D:\Joerg\Research\slides\2025_LeadingHouse_Travel\20241031_CCG_Workshop_Gantt_Chart.xlsx"
    pdf_file = r"D:\Joerg\Research\slides\2025_LeadingHouse_Travel\20241031_CCG_Workshop_Gantt_Chart.pdf"

    # Try COM method first (requires Excel installed)
    if not excel_to_pdf(excel_file, pdf_file):
        print("COM method failed, trying alternative method...")

        # Install required packages if needed
        import subprocess
        import sys

        try:
            from reportlab.lib import colors
        except ImportError:
            print("Installing reportlab...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "reportlab"])

        excel_to_pdf_alternative(excel_file, pdf_file)